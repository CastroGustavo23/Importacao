"""
pipeline_containers.py
======================
Módulo de análise de frete por container — Gocase Importações

Lógica principal:
    - Chave de relacionamento: Nº Container (col AY) + Ref BL (col AE) do Controle PIs
    - Frete de cada processo vem da coluna I ($Frete) do Controle PIs
    - Frete Total do Container = soma de todos os fretes dos processos naquele container
    - Frete Container / TEU = Frete Total Container / TEU do container

Como usar:
    from pipeline_containers import calcular_frete_container, detalhe_container_bl

    df_resumo = calcular_frete_container("Controle de Importações.xlsx")
    df_detalhe = detalhe_container_bl("Controle de Importações.xlsx")

Depende de: pandas, openpyxl
"""

import pandas as pd
import numpy as np
from pathlib import Path

# ──────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────

TEU_MAP = {
    "20'"   : 1.0,
    "40'"   : 2.0,
    "40'HC" : 2.0,
    "LCL"   : 0.0,
}

ABA_CONTROLE   = "Controle PIs"
ABA_CONTAINERS = "Containers"

# Índices das colunas no Controle PIs (base 0)
IDX_REF_BL    = 30   # AE — Ref Trading / BL
IDX_CONTAINER = 50   # AY — Nº Container
IDX_FRETE     = 8    # I  — $ Frete
IDX_VALOR_PI  = 7    # H  — Valor Total PI
IDX_ETD       = 10   # K  — ETD China
IDX_ETA       = 11   # L  — ETA Santos
IDX_MODAL     = 27   # AB — Modal
IDX_SUPPLIER  = 39   # AN — Supplier
IDX_STATUS    = 5    # F  — Status


# ──────────────────────────────────────────────────────────────────
# LEITURA DO CONTROLE PIS
# ──────────────────────────────────────────────────────────────────

def ler_controle_pis(caminho_xlsx: str | Path) -> pd.DataFrame:
    """
    Lê a aba Controle PIs e retorna apenas as linhas com Nº Container
    preenchido na coluna AY.
    """
    print(f"[pipeline] Lendo '{ABA_CONTROLE}'...")

    df = pd.read_excel(caminho_xlsx, sheet_name=ABA_CONTROLE, header=0, dtype=str)

    def to_num(series):
        return pd.to_numeric(
            series.astype(str)
                  .str.replace(",", ".", regex=False)
                  .str.replace(r"[^\d.]", "", regex=True),
            errors="coerce"
        )

    df_sel = pd.DataFrame({
        "ref_bl"       : df.iloc[:, IDX_REF_BL].astype(str).str.strip(),
        "num_container": df.iloc[:, IDX_CONTAINER].astype(str).str.strip().str.upper(),
        "frete"        : to_num(df.iloc[:, IDX_FRETE]),
        "valor_pi"     : to_num(df.iloc[:, IDX_VALOR_PI]),
        "etd"          : pd.to_datetime(df.iloc[:, IDX_ETD],  dayfirst=True, errors="coerce"),
        "eta"          : pd.to_datetime(df.iloc[:, IDX_ETA],  dayfirst=True, errors="coerce"),
        "modal"        : df.iloc[:, IDX_MODAL].astype(str).str.strip(),
        "supplier"     : df.iloc[:, IDX_SUPPLIER].astype(str).str.strip(),
        "status"       : df.iloc[:, IDX_STATUS].astype(str).str.strip(),
    })

    # Filtrar apenas linhas com container válido
    df_sel = df_sel[
        df_sel["num_container"].notna() &
        ~df_sel["num_container"].isin(["", "NAN", "nan", "None"])
    ].copy()

    df_sel["lead_time"] = (df_sel["eta"] - df_sel["etd"]).dt.days

    print(f"  → {len(df_sel)} processos com container preenchido")
    return df_sel


# ──────────────────────────────────────────────────────────────────
# LEITURA DO TIPO DE CADA CONTAINER (aba Containers)
# ──────────────────────────────────────────────────────────────────

def ler_tipo_containers(caminho_xlsx: str | Path) -> pd.DataFrame:
    """
    Lê col A (Nº Container) e col C (Tipo) da aba Containers.
    Retorna 1 linha por container com tipo e TEU.
    """
    print(f"[pipeline] Lendo tipos da aba '{ABA_CONTAINERS}'...")

    try:
        df = pd.read_excel(caminho_xlsx, sheet_name=ABA_CONTAINERS, header=0, dtype=str)
    except Exception as e:
        print(f"  ⚠ Aba '{ABA_CONTAINERS}' não encontrada: {e}")
        return pd.DataFrame(columns=["num_container", "tipo", "teu"])

    df_cont = pd.DataFrame({
        "num_container": df.iloc[:, 0].astype(str).str.strip().str.upper(),
        "tipo"         : df.iloc[:, 2].astype(str).str.strip(),
    })

    df_cont = df_cont[
        df_cont["num_container"].notna() &
        ~df_cont["num_container"].isin(["", "NAN", "nan", "None", "NÃO INFORMADO"])
    ].copy()

    df_cont["teu"] = df_cont["tipo"].map(TEU_MAP).fillna(0)

    # Agregar por container: pegar o tipo com maior TEU (evita pegar NaN ou vazio)
    df_cont = df_cont.sort_values("teu", ascending=False)
    df_cont = df_cont.drop_duplicates(subset="num_container", keep="first").copy()

    print(f"  → {len(df_cont)} containers únicos com tipo definido")
    return df_cont


# ──────────────────────────────────────────────────────────────────
# VISÃO 1: 1 LINHA POR CONTAINER
# ──────────────────────────────────────────────────────────────────

def calcular_frete_container(caminho_xlsx: str | Path) -> pd.DataFrame:
    """
    Agrega todos os processos (BLs) de cada container e calcula:
        - Frete Total do Container (soma dos fretes de todos os processos)
        - Frete Container / TEU
        - % Frete / Valor PI
        - Lead time médio
        - Lista de BLs e fornecedores envolvidos

    Retorna 1 linha por container.
    """
    df_pis  = ler_controle_pis(caminho_xlsx)
    df_tipo = ler_tipo_containers(caminho_xlsx)

    if df_pis.empty:
        return pd.DataFrame()

    grp = df_pis.groupby("num_container")

    resumo = grp.agg(
        qtd_processos    = ("ref_bl",      "count"),
        bls              = ("ref_bl",      lambda x: " | ".join(sorted(x.unique()))),
        suppliers        = ("supplier",    lambda x: " | ".join(sorted(x.dropna().unique()))),
        etd_embarque     = ("etd",         "min"),
        eta_chegada      = ("eta",         "max"),
        frete_total_cont = ("frete",       "sum"),
        valor_pi_total   = ("valor_pi",    "sum"),
        lead_time_medio  = ("lead_time",   "mean"),
    ).reset_index()

    # Merge com tipo/TEU — left join para manter TODOS os containers de df_pis
    if not df_tipo.empty:
        # Normalizar chave antes do merge
        resumo["num_container"] = resumo["num_container"].astype(str).str.strip().str.upper()
        df_tipo["num_container"] = df_tipo["num_container"].astype(str).str.strip().str.upper()
        resumo = resumo.merge(df_tipo, on="num_container", how="left")
    else:
        resumo["tipo"] = ""
        resumo["teu"]  = np.nan

    # Frete / TEU
    resumo["frete_por_teu"] = np.where(
        (resumo["teu"] > 0) & resumo["frete_total_cont"].notna(),
        resumo["frete_total_cont"] / resumo["teu"],
        np.nan
    )

    # % Frete / Valor PI
    resumo["pct_frete_pi"] = np.where(
        (resumo["valor_pi_total"] > 0) & resumo["frete_total_cont"].notna(),
        resumo["frete_total_cont"] / resumo["valor_pi_total"],
        np.nan
    )

    cols = [
        "num_container", "tipo", "teu",
        "qtd_processos", "bls", "suppliers",
        "etd_embarque", "eta_chegada", "lead_time_medio",
        "frete_total_cont", "valor_pi_total",
        "frete_por_teu", "pct_frete_pi",
    ]
    resumo = resumo[[c for c in cols if c in resumo.columns]]
    resumo = resumo.sort_values("frete_total_cont", ascending=False).reset_index(drop=True)

    _log_resumo(resumo)
    return resumo


# ──────────────────────────────────────────────────────────────────
# VISÃO 2: 1 LINHA POR CONTAINER + BL
# ──────────────────────────────────────────────────────────────────

def detalhe_container_bl(caminho_xlsx: str | Path) -> pd.DataFrame:
    """
    Retorna 1 linha por combinação Container + BL com:
        - Frete deste processo
        - Frete Total do Container (soma de todos os processos)
        - Frete Container / TEU
        - Valor PI deste processo
        - % deste processo no frete total do container
        - % Frete / Valor PI

    Útil para popular a aba Containers da planilha.
    """
    df_pis  = ler_controle_pis(caminho_xlsx)
    df_tipo = ler_tipo_containers(caminho_xlsx)

    if df_pis.empty:
        return pd.DataFrame()

    # Frete total e PI total por container
    agg = df_pis.groupby("num_container").agg(
        frete_total_cont   = ("frete",    "sum"),
        valor_pi_total_cont= ("valor_pi", "sum"),
    ).reset_index()

    df = df_pis.merge(agg, on="num_container", how="left")

    if not df_tipo.empty:
        df = df.merge(df_tipo, on="num_container", how="left")
    else:
        df["tipo"] = ""
        df["teu"]  = np.nan

    # Frete do container / TEU
    df["frete_cont_por_teu"] = np.where(
        (df["teu"] > 0) & df["frete_total_cont"].notna(),
        df["frete_total_cont"] / df["teu"],
        np.nan
    )

    # % deste processo no frete total do container
    df["pct_processo_no_cont"] = np.where(
        (df["frete_total_cont"] > 0) & df["frete"].notna(),
        df["frete"] / df["frete_total_cont"],
        np.nan
    )

    # % frete / PI deste processo
    df["pct_frete_pi"] = np.where(
        (df["valor_pi"] > 0) & df["frete"].notna(),
        df["frete"] / df["valor_pi"],
        np.nan
    )

    cols = [
        "num_container", "tipo", "teu",
        "ref_bl", "supplier", "etd", "eta", "modal", "status",
        "frete",                # frete deste processo (BL)
        "frete_total_cont",     # frete total de todos os BLs do container
        "frete_cont_por_teu",   # frete do container / TEU ← métrica principal
        "valor_pi",             # valor PI deste processo
        "valor_pi_total_cont",  # soma PI de todos os processos do container
        "pct_processo_no_cont", # % deste BL no frete total do container
        "pct_frete_pi",         # % frete / valor PI deste processo
    ]
    df = df[[c for c in cols if c in df.columns]]
    df = df.sort_values(["num_container", "ref_bl"]).reset_index(drop=True)

    print(f"  → {len(df)} linhas (Container + BL)")
    return df


# ──────────────────────────────────────────────────────────────────
# EXPORTAR PARA EXCEL
# ──────────────────────────────────────────────────────────────────

def exportar_para_excel(df: pd.DataFrame, caminho_saida: str | Path, nome_aba: str = "Frete por Container"):
    """Exporta DataFrame para Excel com formatação básica."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    s = Side(style="thin", color="CCCCCC")
    brd = Border(left=s, right=s, top=s, bottom=s)

    AZUL  = "1B3A6B"
    CINZA = "F5F5F5"
    VERDE = "D6F0E0"

    cols_usd = {"frete","frete_total_cont","frete_cont_por_teu","frete_por_teu","valor_pi","valor_pi_total_cont","valor_pi_total"}
    cols_pct = {"pct_frete_pi","pct_processo_no_cont","pct_frete_processo"}

    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col.upper().replace("_", " "))
        cell.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        bg = CINZA if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, (col, val) in enumerate(zip(df.columns, row), 1):
            v = None if (isinstance(val, float) and pd.isna(val)) else val
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=VERDE if col in cols_pct else bg)
            cell.alignment = Alignment(horizontal="right" if col in cols_usd | cols_pct else "left", vertical="center")
            cell.border    = brd
            if col in cols_usd: cell.number_format = '"$"#,##0.00'
            elif col in cols_pct: cell.number_format = '0.0%'

    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(w + 4, 42)

    wb.save(caminho_saida)
    print(f"  → Exportado: {caminho_saida} (aba '{nome_aba}')")


# ──────────────────────────────────────────────────────────────────
# LOG
# ──────────────────────────────────────────────────────────────────

def _log_resumo(df: pd.DataFrame):
    print(f"\n── Resumo Frete por Container ──")
    print(f"  Containers únicos  : {len(df)}")
    if "frete_total_cont" in df.columns:
        print(f"  Frete total geral  : ${df['frete_total_cont'].sum():,.2f}")
        print(f"  Frete médio/cont   : ${df['frete_total_cont'].mean():,.2f}")
    if "frete_por_teu" in df.columns and df["frete_por_teu"].notna().any():
        print(f"  Frete médio/TEU    : ${df['frete_por_teu'].mean():,.2f}")
    if "pct_frete_pi" in df.columns and df["pct_frete_pi"].notna().any():
        print(f"  % Frete/PI médio   : {df['pct_frete_pi'].mean()*100:.1f}%")
    if "tipo" in df.columns:
        print(f"  Por tipo           : {df['tipo'].value_counts().to_dict()}")


# ──────────────────────────────────────────────────────────────────
# MAIN — TESTE ISOLADO
# ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    from pathlib import Path

    caminho = sys.argv[1] if len(sys.argv) > 1 else "Controle de Importações.xlsx"
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    print("\n=== VISÃO 1: 1 LINHA POR CONTAINER ===")
    df_resumo = calcular_frete_container(caminho)
    if not df_resumo.empty:
        # CSV para o dashboard
        csv_cont = output_dir / "dim_containers.csv"
        df_resumo.to_csv(csv_cont, sep=";", decimal=",", index=False)
        print(f"  → Salvo: {csv_cont} ({len(df_resumo)} containers)")
        # Excel opcional
        try:
            exportar_para_excel(df_resumo, "frete_por_container.xlsx", "Por Container")
        except Exception as e:
            print(f"  ⚠️  Excel não salvo (feche o arquivo se estiver aberto): {e}")

    print("\n=== VISÃO 2: DETALHE CONTAINER x BL ===")
    df_det = detalhe_container_bl(caminho)
    if not df_det.empty:
        # CSV para o dashboard
        csv_det = output_dir / "dim_containers_det.csv"
        df_det.to_csv(csv_det, sep=";", decimal=",", index=False)
        print(f"  → Salvo: {csv_det} ({len(df_det)} linhas)")
        # Excel opcional
        try:
            exportar_para_excel(df_det, "frete_por_container_detalhe.xlsx", "Container x BL")
        except Exception as e:
            print(f"  ⚠️  Excel não salvo (feche o arquivo se estiver aberto): {e}")

    print("\n✅ CSVs gerados com sucesso em /output")